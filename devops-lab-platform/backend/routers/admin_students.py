import csv
import io
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from sqlalchemy.orm import Session
from typing import List, Optional
from schemas import StudentCreate, StudentUpdate, StudentListOut, StudentOut, ResetPassword, PointsUpdate, PlanUpdate
from database import get_db
import models
from dependencies import require_admin
from auth import get_password_hash
from email_service import send_welcome_email_task

router = APIRouter()

@router.get("", response_model=List[StudentListOut])
def get_students(group_id: Optional[str] = None, db: Session = Depends(get_db), admin=Depends(require_admin)):
    query = db.query(models.Student)
    if group_id:
        query = query.filter(models.Student.group_id == group_id)
    students = query.all()
    res = []
    for s in students:
        s_out = StudentListOut.model_validate(s)
        if s.group:
            s_out.group_name = s.group.name
        # Calculate labs_done and total_score
        subs = db.query(models.Submission).filter_by(student_id=s.id, passed=True).all()
        scores = {}
        for sub in subs:
            scores[sub.lab_id] = max(scores.get(sub.lab_id, 0), sub.score)
        s_out.labs_done = len(scores)
        s_out.total_score = sum(scores.values())
        res.append(s_out)
    return res

@router.post("", response_model=StudentOut)
def create_student(data: StudentCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db), admin=Depends(require_admin)):
    try:
        if db.query(models.Student).filter(models.Student.email == data.email).first():
            raise HTTPException(400, "Email already exists")
        s = models.Student(
            full_name=data.full_name,
            email=data.email,
            password_hash=get_password_hash(data.password),
            group_id=data.group_id
        )
        db.add(s)
        db.commit()
        db.refresh(s)
        
        background_tasks.add_task(send_welcome_email_task, s.email, s.full_name)
        return s
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"DEBUG ERROR: {str(e)}")

@router.get("/{student_id}", response_model=StudentOut)
def get_student(student_id: str, db: Session = Depends(get_db), admin=Depends(require_admin)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s: raise HTTPException(404)
    return s

@router.put("/{student_id}", response_model=StudentOut)
def update_student(student_id: str, data: StudentUpdate, db: Session = Depends(get_db), admin=Depends(require_admin)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s: raise HTTPException(404)
    s.full_name = data.full_name
    s.email = data.email
    s.group_id = data.group_id
    s.is_active = data.is_active
    db.commit()
    db.refresh(s)
    return s

@router.delete("/{student_id}")
def delete_student(student_id: str, db: Session = Depends(get_db), admin=Depends(require_admin)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s: raise HTTPException(404)
    db.delete(s)
    db.commit()
    return {"message": "deleted"}

@router.post("/{student_id}/reset-password")
def reset_password(student_id: str, data: ResetPassword, db: Session = Depends(get_db), admin=Depends(require_admin)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s: raise HTTPException(404)
    s.password_hash = get_password_hash(data.password)
    db.commit()
    return {"message": "password reset"}

@router.post("/bulk")
async def bulk_create_students(background_tasks: BackgroundTasks, file: UploadFile = File(...), db: Session = Depends(get_db), admin=Depends(require_admin)):
    contents = await file.read()
    filename = file.filename or ''
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        headers = [str(cell.value).strip() for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
    else:
        reader = csv.DictReader(io.StringIO(contents.decode('utf-8')))
        rows = list(reader)

    created = 0
    students_to_email = []
    for row in rows:
        email = row.get("email")
        if not email or db.query(models.Student).filter_by(email=email).first():
            continue
        group_name = row.get("group_name")
        group_id = None
        if group_name:
            g = db.query(models.Group).filter_by(name=group_name).first()
            if g: group_id = g.id
        s = models.Student(
            full_name=row.get("full_name", ""),
            email=email,
            password_hash=get_password_hash(row.get("password", "changeme")),
            group_id=group_id
        )
        db.add(s)
        created += 1
        students_to_email.append((s.email, s.full_name))
    db.commit()
    
    for email_addr, name in students_to_email:
        background_tasks.add_task(send_welcome_email_task, email_addr, name)
        
    return {"created": created}

@router.patch("/{student_id}/points")
def update_student_points(student_id: str, data: PointsUpdate, db: Session = Depends(get_db), admin=Depends(require_admin)):
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s:
        raise HTTPException(404, "Student not found")
    old_xp = s.xp
    s.xp = data.xp
    # Recalculate level: 1 level per 500 XP
    s.level = max(1, data.xp // 500 + 1)
    db.commit()
    db.refresh(s)
    return {"student_id": student_id, "xp": s.xp, "level": s.level, "old_xp": old_xp}

@router.patch("/{student_id}/plan")
def update_student_plan(student_id: str, data: PlanUpdate, db: Session = Depends(get_db), admin=Depends(require_admin)):
    valid_plans = ["free", "pro", "enterprise"]
    if data.plan not in valid_plans:
        raise HTTPException(400, f"Invalid plan. Must be one of: {valid_plans}")
    s = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not s:
        raise HTTPException(404, "Student not found")
    s.subscription_tier = data.plan
    db.commit()
    return {"student_id": student_id, "plan": s.subscription_tier}
